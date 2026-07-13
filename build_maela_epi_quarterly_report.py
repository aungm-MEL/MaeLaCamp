from __future__ import annotations

import argparse
from datetime import date, datetime
from pathlib import Path
import re

from openpyxl.styles import PatternFill, numbers
import pandas as pd


SUMMARY_COLUMNS = [
    "Year",
    "period",
    "Organization",
    "Project Name",
    "District (EHO)",
    "Township_EHO",
    "Twp_MIMU",
    "Clinic Name",
    "ALOD_U1",
    "ALOD_U5",
    "ALOD_>5",
    "BCG_U1",
    "BCG_U5",
    "BCG_>5",
    "OPV1_U1",
    "OPV1_U5",
    "OPV1_>5",
    "OPV2_U1",
    "OPV2_U5",
    "OPV2_>5",
    "OPV3_U1",
    "OPV3_U5",
    "OPV3_>5",
    "Penta1_U1",
    "Penta1_U5",
    "Penta1_>5",
    "Penta2_U1",
    "Penta2_U5",
    "Penta2_>5",
    "Penta3_U1",
    "Penta3_U5",
    "Penta3_>5",
    "MMR1_U1",
    "MMR1_U5",
    "MMR1_>5",
    "MMR2_U1",
    "MMR2_U5",
    "MMR2_>5",
    "JE_U1",
    "JE_U5",
    "JE_>5",
    "IPV_U1",
    "IPV_U5",
    "IPV_>5",
    "CD_U1",
    "CD_U5",
    "CD_>5",
    "Td1",
    "Td2",
    "Td At least one dose",
]

INDICATOR_COLUMNS = [
    "Period",
    "Organization",
    "Project Name",
    "indicator",
    "Q1 Target",
    "Q1 U1 Male",
    "Q1 U1 Female",
    "Q1 1-5 Male",
    "Q1 1-5 Female",
    "Q1 Total",
    "Q2 Target",
    "Q2 U1 Male",
    "Q2 U1 Female",
    "Q2 1-5 Male",
    "Q2 1-5 Female",
    "Q2 Total",
    "Q3 Target",
    "Q3 U1 Male",
    "Q3 U1 Female",
    "Q3 1-5 Male",
    "Q3 1-5 Female",
    "Q3 Total",
    "Q4 Target",
    "Q4 U1 Male",
    "Q4 U1 Female",
    "Q4 1-5 Male",
    "Q4 1-5 Female",
    "Q4 Total",
]

INDICATOR_NAMES = [
    "Penta3 under 1-yr-old",
    "MMR1 under 1-yr-old",
    "Penta1 under 5-yr-old",
    "Penta3 under 5-yr-old",
    "MMR1 under 5-yr-old",
    "MMR2 under 5-yr-old",
    "Full dose under 5-yr-old",
    "At least one dose under 5-yr-old",
    "Td ALOD",
    "Td Two Doses",
]


DEFAULT_INPUT_PATH = Path(__file__).resolve().parent / "MaeLa_EPI_report.xls"
DEFAULT_OUTPUT_PATH = Path(__file__).resolve().parent / "MaeLa_Camp_EPI_Quarterly_Report.xlsx"

DATE_COLUMNS_TO_CONVERT = [
    "patient_birthday",
    "bcg_adbirth",
    "hb1_adbirth",
    "VitA_adbirth",
    "dtp1",
    "ipv1",
    "rota1",
    "vita2months",
    "dtp2",
    "ipv2",
    "rota2",
    "vita4months",
    "dtp3",
    "opv3",
    "rota3",
    "mmr1",
    "vita9months",
    "laje112months",
    "dtp4",
    "opv4",
    "mrr2",
    "vita18months",
    "laje2",
    "dtp5",
    "opv5",
]

FIRST_VISIT_SOURCE_COLUMNS = [
    "bcg_adbirth",
    "dtp1",
    "ipv1",
    "rota1",
    "dtp2",
    "ipv2",
    "rota2",
    "dtp3",
    "opv3",
    "rota3",
    "mmr1",
    "laje112months",
    "dtp4",
    "opv4",
    "mrr2",
    "laje2",
    "dtp5",
    "opv5",
]

LONG_FORM_VACCINE_COLUMNS = [
    "bcg_adbirth",
    "dtp1",
    "ipv1",
    "rota1",
    "dtp2",
    "dtp3",
    "ipv2",
    "rota2",
    "opv3",
    "rota3",
    "mmr1",
    "laje112months",
    "dtp4",
    "opv4",
    "mrr2",
    "laje2",
    "dtp5",
    "opv5",
]

DROP_COLUMNS = [
    "hb1_adbirth",
    "VitA_adbirth",
    "vita2months",
    "vita4months",
    "vita9months",
    "vita18months",
    "patient_house",
    "patient_road",
    "mother_father_name",
    "telephone",
]

ALLOWED_SEX_VALUES = {"male", "female"}
_DATE_LEADING_YEAR_RE = re.compile(r"^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})")
_DATE_TRAILING_YEAR_RE = re.compile(r"^(\d{1,2})[-/.](\d{1,2})[-/.](\d{4})")


def normalize_colname(name: object) -> str:
    return str(name).strip().lower()


def build_column_map(frame: pd.DataFrame) -> dict[str, str]:
    return {normalize_colname(col): str(col) for col in frame.columns}


def thai_to_gregorian(value: object) -> object:
    if pd.isna(value):
        return pd.NA

    if isinstance(value, pd.Timestamp):
        value = value.to_pydatetime()

    if isinstance(value, (datetime, date)):
        year = int(value.year)
        if year >= 2400:
            year -= 543
        try:
            return datetime(year, int(value.month), int(value.day))
        except ValueError:
            return value

    text = str(value).strip()
    if not text:
        return pd.NA

    for pattern, y_idx, m_idx, d_idx in (
        (_DATE_LEADING_YEAR_RE, 1, 2, 3),
        (_DATE_TRAILING_YEAR_RE, 3, 2, 1),
    ):
        match = pattern.match(text)
        if not match:
            continue

        year = int(match.group(y_idx))
        month = int(match.group(m_idx))
        day = int(match.group(d_idx))

        if year >= 2400:
            year -= 543

        try:
            return datetime(year, month, day)
        except ValueError:
            return value

    return value


def check_patient_hn_duplicates(frame: pd.DataFrame, column_name: str) -> tuple[int, list[str]]:
    values = frame[column_name].astype("string").str.strip()
    values = values[values.notna() & values.ne("")]

    duplicated_values = values[values.duplicated(keep=False)]
    duplicate_hns = sorted(duplicated_values.dropna().unique().tolist())
    duplicate_row_count = int(duplicated_values.shape[0])
    return duplicate_row_count, duplicate_hns


def check_sex_description(frame: pd.DataFrame, column_name: str) -> list[str]:
    values = frame[column_name].astype("string").str.strip()
    values = values[values.notna() & values.ne("")]

    invalid_values = sorted(
        {
            str(value)
            for value in values
            if str(value).strip().lower() not in ALLOWED_SEX_VALUES
        }
    )
    return invalid_values


def apply_date_conversion(frame: pd.DataFrame) -> list[str]:
    normalized_map = build_column_map(frame)
    converted_columns: list[str] = []

    for requested_column in DATE_COLUMNS_TO_CONVERT:
        actual_column = normalized_map.get(normalize_colname(requested_column))
        if not actual_column:
            continue

        frame[actual_column] = frame[actual_column].apply(thai_to_gregorian)
        frame[actual_column] = pd.to_datetime(frame[actual_column], errors="coerce")
        converted_columns.append(actual_column)

    return converted_columns


def find_birthdate_after_other_dates(
    frame: pd.DataFrame,
    birth_column: str,
    compare_columns: list[str],
    patient_hn_column: str | None,
) -> tuple[int, list[str], list[str], list[int]]:
    compare_actual_columns = [col for col in compare_columns if col != birth_column]
    if not compare_actual_columns:
        return 0, [], [], []

    birth_dates = pd.to_datetime(frame[birth_column], errors="coerce")
    has_issue_mask = pd.Series(False, index=frame.index)

    for compare_column in compare_actual_columns:
        compare_dates = pd.to_datetime(frame[compare_column], errors="coerce")
        issue_mask = birth_dates.notna() & compare_dates.notna() & birth_dates.gt(compare_dates)
        has_issue_mask = has_issue_mask | issue_mask

    issue_count = int(has_issue_mask.sum())
    issue_indices = has_issue_mask[has_issue_mask].index.tolist()

    duplicate_hns: list[str] = []
    if issue_count > 0 and patient_hn_column and patient_hn_column in frame.columns:
        hns = (
            frame.loc[has_issue_mask, patient_hn_column]
            .astype("string")
            .str.strip()
        )
        hns = hns[hns.notna() & hns.ne("")]
        duplicate_hns = sorted(hns.unique().tolist())

    return issue_count, duplicate_hns, compare_actual_columns, issue_indices


def completed_months_between(start_value: object, end_value: object) -> object:
    start = pd.to_datetime(start_value, errors="coerce")
    end = pd.to_datetime(end_value, errors="coerce")

    if pd.isna(start) or pd.isna(end):
        return pd.NA

    if end < start:
        return "error age"

    month_diff = (end.year - start.year) * 12 + (end.month - start.month)
    if end.day < start.day:
        month_diff -= 1
    return int(month_diff)


def add_first_visit_columns(frame: pd.DataFrame, normalized_map: dict[str, str]) -> tuple[str | None, str | None, list[str]]:
    birth_column = normalized_map.get("patient_birthday")
    source_columns: list[str] = []

    for requested_column in FIRST_VISIT_SOURCE_COLUMNS:
        actual_column = normalized_map.get(normalize_colname(requested_column))
        if actual_column:
            source_columns.append(actual_column)

    if not birth_column or not source_columns:
        return None, None, source_columns

    first_visit_series = frame[source_columns].min(axis=1, skipna=True)
    frame["first_visit_date"] = pd.to_datetime(first_visit_series, errors="coerce")

    frame["first_visit_age"] = [
        completed_months_between(birth, first_visit)
        for birth, first_visit in zip(frame[birth_column], frame["first_visit_date"])
    ]

    return "first_visit_date", "first_visit_age", source_columns


def build_long_form_sheet(frame: pd.DataFrame, normalized_map: dict[str, str]) -> pd.DataFrame:
    required_base_columns = ["child_name", "patient_hn", "sex_description", "patient_birthday"]
    optional_base_columns = ["first_visit_date", "first_visit_age"]
    long_form_id_columns = required_base_columns + optional_base_columns

    base_actual: list[str] = []
    base_rename: dict[str, str] = {}
    for requested in required_base_columns:
        actual = normalized_map.get(requested)
        if not actual:
            raise KeyError(f"Column '{requested}' was not found for long_form sheet.")
        base_actual.append(actual)
        base_rename[actual] = requested

    for requested in optional_base_columns:
        actual = normalized_map.get(requested)
        if actual:
            base_actual.append(actual)
            base_rename[actual] = requested

    vaccine_actual: list[str] = []
    vaccine_rename: dict[str, str] = {}
    for requested in LONG_FORM_VACCINE_COLUMNS:
        actual = normalized_map.get(normalize_colname(requested))
        if actual:
            vaccine_actual.append(actual)
            vaccine_rename[actual] = requested

    if not vaccine_actual:
        raise KeyError("None of the requested vaccine columns were found for long_form sheet.")

    selected = frame[base_actual + vaccine_actual].copy()
    selected.rename(columns=base_rename, inplace=True)

    for optional_column in optional_base_columns:
        if optional_column not in selected.columns:
            selected[optional_column] = pd.NA

    long_form = selected.melt(
        id_vars=long_form_id_columns,
        value_vars=vaccine_actual,
        var_name="vaccine",
        value_name="vaccine_date",
    )

    long_form["vaccine"] = long_form["vaccine"].map(vaccine_rename).fillna(long_form["vaccine"])
    long_form["patient_birthday"] = pd.to_datetime(long_form["patient_birthday"], errors="coerce")
    long_form["first_visit_date"] = pd.to_datetime(long_form["first_visit_date"], errors="coerce")
    long_form["vaccine_date"] = pd.to_datetime(long_form["vaccine_date"], errors="coerce")
    long_form = long_form[long_form["vaccine_date"].notna()].reset_index(drop=True)

    age_at_dose_values: list[object] = []
    error_mask = pd.Series(False, index=long_form.index)
    for idx, (birth, vaccine_date) in enumerate(zip(long_form["patient_birthday"], long_form["vaccine_date"])):
        age_value = completed_months_between(birth, vaccine_date)
        if age_value == "error age":
            age_at_dose_values.append("error")
            error_mask.iloc[idx] = True
        else:
            age_at_dose_values.append(age_value)

    long_form["age_at_dose"] = age_at_dose_values

    error_hns = (
        long_form.loc[error_mask, "patient_hn"]
        .astype("string")
        .str.strip()
    )
    error_hns = error_hns[error_hns.notna() & error_hns.ne("")]
    age_at_dose_error_hns = sorted(error_hns.unique().tolist())
    long_form.attrs["age_at_dose_error_count"] = int(error_mask.sum())
    long_form.attrs["age_at_dose_error_hns"] = age_at_dose_error_hns

    return long_form


def map_vaccine_group_from_maela(vaccine_value: object) -> str | None:
    vaccine = str(vaccine_value).strip().lower()
    if vaccine == "bcg_adbirth":
        return "BCG"
    if vaccine == "dtp1":
        return "Penta1"
    if vaccine == "dtp2":
        return "Penta2"
    if vaccine == "dtp3":
        return "Penta3"
    if vaccine == "opv1":
        return "OPV1"
    if vaccine == "opv2":
        return "OPV2"
    if vaccine == "opv3":
        return "OPV3"
    if vaccine == "opv":
        return "OPV"
    if vaccine == "opv4":
        return None
    if vaccine == "opv5":
        return None
    if vaccine == "ipv1":
        return "IPV"
    if vaccine == "ipv2":
        return None
    if vaccine == "mmr1":
        return "MMR1"
    if vaccine == "mrr2":
        return "MMR2"
    if vaccine in {"laje112months", "laje2"}:
        return "JE"
    return None


def create_summary_sheet_from_long_form(long_form: pd.DataFrame) -> pd.DataFrame:
    if long_form.empty:
        print("WARNING: Summary sheet generated with no rows because long_form is empty.")
        return pd.DataFrame(columns=SUMMARY_COLUMNS)

    work = long_form.copy()
    work["vaccine_date"] = pd.to_datetime(work["vaccine_date"], errors="coerce")
    work = work.dropna(subset=["vaccine_date", "patient_hn"])
    if work.empty:
        print("WARNING: Summary sheet generated with no rows because required values are missing.")
        return pd.DataFrame(columns=SUMMARY_COLUMNS)

    work["beneficiary_id"] = work["patient_hn"].astype(str).str.strip()
    work = work[work["beneficiary_id"] != ""]

    # Requested exclusion for Summary: remove OPV4, OPV5, and IPV2.
    excluded_summary_vaccines = {"opv4", "opv5", "ipv2"}
    vaccine_key = work["vaccine"].astype(str).str.strip().str.lower()
    work = work.loc[~vaccine_key.isin(excluded_summary_vaccines)].copy()

    work["Year"] = work["vaccine_date"].dt.year.astype("Int64")
    work["period"] = "Q" + work["vaccine_date"].dt.quarter.astype("Int64").astype(str) + "_" + work["Year"].astype(str)

    work["vaccine_group"] = work["vaccine"].map(map_vaccine_group_from_maela)
    # Antigen dose columns must be calculated by age_at_dose.
    work["dose_age_months"] = pd.to_numeric(work["age_at_dose"], errors="coerce")
    # ALOD columns must be calculated by first_visit_age.
    work["first_visit_age_num"] = pd.to_numeric(work.get("first_visit_age"), errors="coerce")

    summary_rows: list[dict[str, object]] = []

    u1_rules: dict[str, tuple[int, int]] = {
        "BCG": (0, 11),
        "Penta1": (2, 11),
        "Penta2": (3, 11),
        "Penta3": (4, 11),
        "MMR1": (9, 11),
    }
    default_u1_rule = (0, 11)
    u5_rule = (12, 59)

    for period_value in sorted(work["period"].dropna().unique()):
        period_frame = work[work["period"] == period_value]
        year_value = int(period_frame["Year"].dropna().iloc[0])

        row: dict[str, object] = {
            "Year": year_value,
            "period": period_value,
            "Organization": "MaeLa Camp",
            "Project Name": "REACH-KK",
            "District (EHO)": "",
            "Township_EHO": "",
            "Twp_MIMU": "",
            "Clinic Name": "MaeLa Camp",
        }

        def unique_count_by_age(min_age: int, max_age: int, vaccine_group: str | None = None) -> int:
            subset = period_frame[period_frame["dose_age_months"].between(min_age, max_age, inclusive="both")]
            if vaccine_group is not None:
                subset = subset[subset["vaccine_group"] == vaccine_group]
            return int(subset["beneficiary_id"].nunique())

        def unique_count_over_5(vaccine_group: str | None = None) -> int:
            subset = period_frame[period_frame["dose_age_months"] >= 60]
            if vaccine_group is not None:
                subset = subset[subset["vaccine_group"] == vaccine_group]
            return int(subset["beneficiary_id"].nunique())

        alod_period = period_frame.dropna(subset=["first_visit_age_num", "beneficiary_id"])
        alod_period = alod_period.sort_values("vaccine_date").drop_duplicates(subset=["beneficiary_id"], keep="first")
        row["ALOD_U1"] = int(
            alod_period[alod_period["first_visit_age_num"].between(0, 11, inclusive="both")]["beneficiary_id"].nunique()
        )
        row["ALOD_U5"] = int(
            alod_period[alod_period["first_visit_age_num"].between(12, 59, inclusive="both")]["beneficiary_id"].nunique()
        )
        row["ALOD_>5"] = int(
            alod_period[alod_period["first_visit_age_num"] >= 60]["beneficiary_id"].nunique()
        )

        for antigen in ["BCG", "OPV1", "OPV2", "OPV3", "Penta1", "Penta2", "Penta3", "MMR1", "MMR2", "JE", "IPV", "CD"]:
            u1_min, u1_max = u1_rules.get(antigen, default_u1_rule)
            row[f"{antigen}_U1"] = unique_count_by_age(u1_min, u1_max, antigen)
            row[f"{antigen}_U5"] = unique_count_by_age(u5_rule[0], u5_rule[1], antigen)
            row[f"{antigen}_>5"] = unique_count_over_5(antigen)

        row["Td1"] = int(period_frame[period_frame["vaccine_group"] == "Td1"]["beneficiary_id"].nunique())
        row["Td2"] = int(period_frame[period_frame["vaccine_group"] == "Td2"]["beneficiary_id"].nunique())
        td_any_mask = period_frame["vaccine_group"].isin(["Td1", "Td2", "TdAny"])
        row["Td At least one dose"] = int(period_frame[td_any_mask]["beneficiary_id"].nunique())

        summary_rows.append(row)

    summary = pd.DataFrame(summary_rows)
    summary = summary.reindex(columns=SUMMARY_COLUMNS, fill_value=0)
    print(f"INFO: Created Summary sheet with {len(summary)} period row(s)")
    return summary


def create_indicator_sheet_from_long_form(long_form: pd.DataFrame) -> pd.DataFrame:
    if long_form.empty:
        print("WARNING: Indicator sheet could not be generated because long_form is empty.")
        return pd.DataFrame(columns=INDICATOR_COLUMNS)

    work = long_form.copy()
    work["vaccine_date"] = pd.to_datetime(work["vaccine_date"], errors="coerce")
    work = work.dropna(subset=["vaccine_date", "patient_hn"])
    if work.empty:
        print("WARNING: Indicator sheet generated with no rows because required values are missing.")
        return pd.DataFrame(columns=INDICATOR_COLUMNS)

    work["beneficiary_id"] = work["patient_hn"].astype(str).str.strip()
    work = work[work["beneficiary_id"] != ""]

    # Keep exclusions aligned with Summary calculations.
    excluded_vaccines = {"opv4", "opv5", "ipv2"}
    vaccine_key = work["vaccine"].astype(str).str.strip().str.lower()
    work = work.loc[~vaccine_key.isin(excluded_vaccines)].copy()

    work["Year"] = work["vaccine_date"].dt.year.astype("Int64")
    work["Quarter"] = work["vaccine_date"].dt.quarter.astype("Int64")
    work["vaccine_group"] = work["vaccine"].map(map_vaccine_group_from_maela)
    work["age_months"] = pd.to_numeric(work.get("age_at_dose"), errors="coerce")

    sex_norm = work["sex_description"].astype(str).str.strip().str.lower()
    work["sex_norm"] = sex_norm.map({"male": "M", "female": "F", "m": "M", "f": "F"}).fillna("")

    u1_rules: dict[str, tuple[int, int]] = {
        "BCG": (0, 11),
        "Penta1": (2, 11),
        "Penta2": (3, 11),
        "Penta3": (4, 11),
        "MMR1": (9, 11),
    }
    default_u1_rule = (0, 11)
    u5_rule = (12, 59)

    def indicator_mask(indicator: str) -> pd.Series:
        if indicator == "Penta3 under 1-yr-old":
            return work["vaccine_group"] == "Penta3"
        if indicator == "MMR1 under 1-yr-old":
            return work["vaccine_group"] == "MMR1"
        if indicator == "Penta1 under 5-yr-old":
            return work["vaccine_group"] == "Penta1"
        if indicator == "Penta3 under 5-yr-old":
            return work["vaccine_group"] == "Penta3"
        if indicator == "MMR1 under 5-yr-old":
            return work["vaccine_group"] == "MMR1"
        if indicator == "MMR2 under 5-yr-old":
            return work["vaccine_group"] == "MMR2"
        if indicator == "Full dose under 5-yr-old":
            return work["vaccine_group"] == "Penta3"
        if indicator == "At least one dose under 5-yr-old":
            return work["vaccine_group"].notna()
        if indicator == "Td ALOD":
            return work["vaccine_group"].isin(["Td1", "Td2", "TdAny"])
        if indicator == "Td Two Doses":
            return work["vaccine_group"] == "Td2"
        return pd.Series(False, index=work.index)

    def u1_range_for_indicator(indicator: str) -> tuple[int, int]:
        if "Penta3" in indicator:
            return u1_rules["Penta3"]
        if "Penta1" in indicator:
            return u1_rules["Penta1"]
        if "MMR1" in indicator:
            return u1_rules["MMR1"]
        if "MMR2" in indicator:
            return default_u1_rule
        return default_u1_rule

    rows: list[dict[str, object]] = []
    years = sorted(work["Year"].dropna().astype(int).unique())
    for year_value in years:
        for indicator in INDICATOR_NAMES:
            row: dict[str, object] = {
                "Period": year_value,
                "Organization": "MaeLa Camp",
                "Project Name": "REACH-KK",
                "indicator": indicator,
            }

            base_mask = indicator_mask(indicator) & work["Year"].eq(year_value)
            u1_min, u1_max = u1_range_for_indicator(indicator)

            for quarter in [1, 2, 3, 4]:
                quarter_mask = base_mask & work["Quarter"].eq(quarter)

                if indicator in {"Td ALOD", "Td Two Doses"}:
                    td_female_count = int(
                        work.loc[quarter_mask & work["sex_norm"].eq("F"), "beneficiary_id"].nunique()
                    )
                    q_u1_m = 0
                    q_u1_f = 0
                    q_u5_m = 0
                    q_u5_f = td_female_count
                    q_total = q_u5_f
                else:
                    u1_mask = quarter_mask & work["age_months"].between(u1_min, u1_max, inclusive="both")
                    u5_mask = quarter_mask & work["age_months"].between(u5_rule[0], u5_rule[1], inclusive="both")

                    q_u1_m = int(work.loc[u1_mask & work["sex_norm"].eq("M"), "beneficiary_id"].nunique())
                    q_u1_f = int(work.loc[u1_mask & work["sex_norm"].eq("F"), "beneficiary_id"].nunique())

                    if "under 1-yr-old" in indicator.lower():
                        q_u5_m = ""
                        q_u5_f = ""
                        q_total = q_u1_m + q_u1_f
                    else:
                        q_u5_m = int(work.loc[u5_mask & work["sex_norm"].eq("M"), "beneficiary_id"].nunique())
                        q_u5_f = int(work.loc[u5_mask & work["sex_norm"].eq("F"), "beneficiary_id"].nunique())
                        q_total = q_u1_m + q_u1_f + q_u5_m + q_u5_f

                row[f"Q{quarter} Target"] = 0
                row[f"Q{quarter} U1 Male"] = q_u1_m
                row[f"Q{quarter} U1 Female"] = q_u1_f
                row[f"Q{quarter} 1-5 Male"] = q_u5_m
                row[f"Q{quarter} 1-5 Female"] = q_u5_f
                row[f"Q{quarter} Total"] = q_total

            rows.append(row)

    indicator_df = pd.DataFrame(rows).reindex(columns=INDICATOR_COLUMNS, fill_value=0)
    print(f"INFO: Created indicator sheet with {len(indicator_df)} row(s)")
    return indicator_df


def apply_excel_date_format(
    writer: pd.ExcelWriter,
    sheet_name: str,
    frame: pd.DataFrame,
    date_columns: list[str],
) -> None:
    worksheet = writer.sheets[sheet_name]
    date_columns_set = set(date_columns)

    for column_index, column_name in enumerate(frame.columns, start=1):
        if str(column_name) not in date_columns_set:
            continue

        # Row 1 is header; apply formatting to data rows only.
        for row_index in range(2, len(frame) + 2):
            cell = worksheet.cell(row=row_index, column=column_index)
            if cell.value is not None:
                cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2


def apply_birthdate_error_hn_highlight(
    writer: pd.ExcelWriter,
    sheet_name: str,
    frame: pd.DataFrame,
    patient_hn_column: str,
    issue_row_indices: list[int],
) -> None:
    if not issue_row_indices or patient_hn_column not in frame.columns:
        return

    worksheet = writer.sheets[sheet_name]
    hn_column_index = list(frame.columns).index(patient_hn_column) + 1
    red_fill = PatternFill(fill_type="solid", start_color="00FFC7CE", end_color="00FFC7CE")

    for row_idx in issue_row_indices:
        # Data starts at Excel row 2 because row 1 is header.
        excel_row = int(row_idx) + 2
        cell = worksheet.cell(row=excel_row, column=hn_column_index)
        cell.fill = red_fill


def drop_requested_columns(frame: pd.DataFrame) -> list[str]:
    normalized_map = build_column_map(frame)
    columns_to_drop: list[str] = []

    for requested_column in DROP_COLUMNS:
        actual_column = normalized_map.get(normalize_colname(requested_column))
        if actual_column:
            columns_to_drop.append(actual_column)

    if columns_to_drop:
        frame.drop(columns=columns_to_drop, inplace=True)

    return columns_to_drop


def build_validation_report(
    duplicate_row_count: int,
    duplicate_hns: list[str],
    invalid_sex_values: list[str],
    converted_columns: list[str],
    dropped_columns: list[str],
    birthdate_warning_count: int,
    birthdate_warning_hns: list[str],
    birthdate_compare_columns: list[str],
    first_visit_date_col: str | None,
    first_visit_age_col: str | None,
    first_visit_source_columns: list[str],
    age_at_dose_error_count: int,
    age_at_dose_error_hns: list[str],
) -> pd.DataFrame:
    report_rows: list[dict[str, str]] = []

    if duplicate_hns:
        report_rows.append(
            {
                "check": "patient_hn duplicates",
                "status": "WARNING",
                "details": (
                    f"Found {duplicate_row_count} duplicated row(s) in patient_hn. "
                    f"Duplicated HN values: {', '.join(duplicate_hns)}"
                ),
            }
        )
    else:
        report_rows.append(
            {
                "check": "patient_hn duplicates",
                "status": "OK",
                "details": "No HN duplication was found.",
            }
        )

    if invalid_sex_values:
        report_rows.append(
            {
                "check": "Sex_description values",
                "status": "WARNING",
                "details": (
                    "Values other than Male/Female were found in Sex_description: "
                    + ", ".join(invalid_sex_values)
                ),
            }
        )
    else:
        report_rows.append(
            {
                "check": "Sex_description values",
                "status": "OK",
                "details": "Sex_description contains only Male/Female (or empty values).",
            }
        )

    report_rows.append(
        {
            "check": "Thai date conversion",
            "status": "INFO",
            "details": (
                "Converted Thai calendar dates to Gregorian in columns: "
                + (", ".join(converted_columns) if converted_columns else "None of the requested columns were found.")
            ),
        }
    )

    report_rows.append(
        {
            "check": "Column removal",
            "status": "INFO",
            "details": (
                "Removed columns: "
                + (", ".join(dropped_columns) if dropped_columns else "None of the requested columns were found.")
            ),
        }
    )

    if birthdate_warning_count > 0:
        hn_details = ", ".join(birthdate_warning_hns) if birthdate_warning_hns else "HN values unavailable"
        report_rows.append(
            {
                "check": "Birthdate later than visit date",
                "status": "WARNING",
                "details": (
                    f"Found {birthdate_warning_count} row(s) where patient_birthday is later than one or more visit/date columns. "
                    f"Compared columns: {', '.join(birthdate_compare_columns)}. "
                    f"Affected patient_hn: {hn_details}"
                ),
            }
        )
    else:
        report_rows.append(
            {
                "check": "Birthdate later than visit date",
                "status": "OK",
                "details": "No rows found where patient_birthday is later than the compared date columns.",
            }
        )

    if first_visit_date_col and first_visit_age_col:
        report_rows.append(
            {
                "check": "First visit columns",
                "status": "INFO",
                "details": (
                    f"Added columns {first_visit_date_col} and {first_visit_age_col}. "
                    f"first_visit_date is minimum of: {', '.join(first_visit_source_columns)}. "
                    "first_visit_age is completed month difference like DATEDIF(birth, first_visit_date, \"M\")."
                ),
            }
        )
    else:
        report_rows.append(
            {
                "check": "First visit columns",
                "status": "WARNING",
                "details": "Could not add first_visit_date/first_visit_age because patient_birthday or source columns were not found.",
            }
        )

    if age_at_dose_error_count > 0:
        hn_details = ", ".join(age_at_dose_error_hns) if age_at_dose_error_hns else "HN values unavailable"
        report_rows.append(
            {
                "check": "age_at_dose logical errors",
                "status": "WARNING",
                "details": (
                    f"Found {age_at_dose_error_count} long_form row(s) where patient_birthday is later than vaccine_date. "
                    "age_at_dose is set to 'error' for those rows. "
                    f"Affected patient_hn: {hn_details}"
                ),
            }
        )
    else:
        report_rows.append(
            {
                "check": "age_at_dose logical errors",
                "status": "OK",
                "details": "No long_form rows found where patient_birthday is later than vaccine_date.",
            }
        )

    return pd.DataFrame(report_rows)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Build MaeLa quarterly EPI report by validating HN/Sex_description, "
            "converting Thai dates to Gregorian, removing requested columns, and exporting XLSX."
        )
    )
    parser.add_argument(
        "--input",
        default=str(DEFAULT_INPUT_PATH),
        help="Path to input MaeLa_EPI_report.xls file.",
    )
    parser.add_argument(
        "--output",
        default=str(DEFAULT_OUTPUT_PATH),
        help="Path to output MaeLa_Camp_EPI_Quarterly_Report.xlsx file.",
    )
    parser.add_argument(
        "--sheet",
        default=0,
        help="Sheet name or 0-based index to read from the source workbook. Default is first sheet.",
    )
    return parser.parse_args()


def build_maela_report(input_path: Path, output_path: Path, sheet_arg: str | int = 0) -> Path:
    input_path = Path(input_path).resolve()
    output_path = Path(output_path).resolve()

    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")

    # .xls files usually require xlrd; xlsx files use openpyxl.
    engine = "xlrd" if input_path.suffix.lower() == ".xls" else "openpyxl"
    frame = pd.read_excel(input_path, sheet_name=sheet_arg, engine=engine)

    normalized_map = build_column_map(frame)

    patient_hn_column = normalized_map.get("patient_hn")
    if not patient_hn_column:
        raise KeyError("Column 'patient_hn' was not found in the source file.")

    sex_description_column = normalized_map.get("sex_description")
    if not sex_description_column:
        raise KeyError("Column 'Sex_description' was not found in the source file.")

    duplicate_row_count, duplicate_hns = check_patient_hn_duplicates(frame, patient_hn_column)
    invalid_sex_values = check_sex_description(frame, sex_description_column)
    converted_columns = apply_date_conversion(frame)
    normalized_map = build_column_map(frame)

    first_visit_date_col, first_visit_age_col, first_visit_source_columns = add_first_visit_columns(frame, normalized_map)

    birth_column = normalized_map.get("patient_birthday")
    compare_columns = [
        actual for requested in DATE_COLUMNS_TO_CONVERT
        if (actual := normalized_map.get(normalize_colname(requested)))
    ]
    if first_visit_date_col:
        compare_columns.append(first_visit_date_col)

    birthdate_warning_count = 0
    birthdate_warning_hns: list[str] = []
    birthdate_compare_columns: list[str] = []
    birthdate_issue_row_indices: list[int] = []
    if birth_column:
        birthdate_warning_count, birthdate_warning_hns, birthdate_compare_columns, birthdate_issue_row_indices = find_birthdate_after_other_dates(
            frame=frame,
            birth_column=birth_column,
            compare_columns=compare_columns,
            patient_hn_column=patient_hn_column,
        )

    dropped_columns = drop_requested_columns(frame)

    # Build long_form from the same cleaned dataframe that is exported to MaeLa_EPI_Clean.
    clean_export_frame = frame
    clean_normalized_map = build_column_map(clean_export_frame)
    long_form = build_long_form_sheet(clean_export_frame, clean_normalized_map)
    summary_sheet = create_summary_sheet_from_long_form(long_form)
    indicator_sheet = create_indicator_sheet_from_long_form(long_form)
    age_at_dose_error_count = int(long_form.attrs.get("age_at_dose_error_count", 0))
    age_at_dose_error_hns = list(long_form.attrs.get("age_at_dose_error_hns", []))

    validation_report = build_validation_report(
        duplicate_row_count=duplicate_row_count,
        duplicate_hns=duplicate_hns,
        invalid_sex_values=invalid_sex_values,
        converted_columns=converted_columns,
        dropped_columns=dropped_columns,
        birthdate_warning_count=birthdate_warning_count,
        birthdate_warning_hns=birthdate_warning_hns,
        birthdate_compare_columns=birthdate_compare_columns,
        first_visit_date_col=first_visit_date_col,
        first_visit_age_col=first_visit_age_col,
        first_visit_source_columns=first_visit_source_columns,
        age_at_dose_error_count=age_at_dose_error_count,
        age_at_dose_error_hns=age_at_dose_error_hns,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)

    date_format_columns = list(converted_columns)
    if first_visit_date_col:
        date_format_columns.append(first_visit_date_col)

    final_output_path = output_path
    try:
        with pd.ExcelWriter(final_output_path, engine="openpyxl") as writer:
            clean_export_frame.to_excel(writer, sheet_name="MaeLa_EPI_Clean", index=False)
            apply_excel_date_format(writer, "MaeLa_EPI_Clean", clean_export_frame, date_format_columns)
            apply_birthdate_error_hn_highlight(
                writer,
                "MaeLa_EPI_Clean",
                clean_export_frame,
                patient_hn_column,
                birthdate_issue_row_indices,
            )
            long_form.to_excel(writer, sheet_name="long_form", index=False)
            apply_excel_date_format(writer, "long_form", long_form, ["patient_birthday", "first_visit_date", "vaccine_date"])
            summary_sheet.to_excel(writer, sheet_name="Summary", index=False)
            indicator_sheet.to_excel(writer, sheet_name="indicator", index=False)
            validation_report.to_excel(writer, sheet_name="Validation_Report", index=False)
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        final_output_path = output_path.with_name(f"{output_path.stem}_{timestamp}{output_path.suffix}")
        with pd.ExcelWriter(final_output_path, engine="openpyxl") as writer:
            clean_export_frame.to_excel(writer, sheet_name="MaeLa_EPI_Clean", index=False)
            apply_excel_date_format(writer, "MaeLa_EPI_Clean", clean_export_frame, date_format_columns)
            apply_birthdate_error_hn_highlight(
                writer,
                "MaeLa_EPI_Clean",
                clean_export_frame,
                patient_hn_column,
                birthdate_issue_row_indices,
            )
            long_form.to_excel(writer, sheet_name="long_form", index=False)
            apply_excel_date_format(writer, "long_form", long_form, ["patient_birthday", "first_visit_date", "vaccine_date"])
            summary_sheet.to_excel(writer, sheet_name="Summary", index=False)
            indicator_sheet.to_excel(writer, sheet_name="indicator", index=False)
            validation_report.to_excel(writer, sheet_name="Validation_Report", index=False)
        print(
            "[WARNING] Output file was locked, saved to fallback file instead: "
            f"{final_output_path}"
        )

    for _, row in validation_report.iterrows():
        print(f"[{row['status']}] {row['check']}: {row['details']}")
    print(f"[INFO] summary rows: {len(summary_sheet)}")
    print(f"[INFO] indicator rows: {len(indicator_sheet)}")
    print(f"[INFO] long_form rows: {len(long_form)}")
    print(f"Saved output file: {final_output_path}")
    return final_output_path


def main() -> None:
    args = parse_args()
    parsed_sheet: str | int
    parsed_sheet = int(args.sheet) if str(args.sheet).isdigit() else args.sheet
    build_maela_report(
        input_path=Path(args.input),
        output_path=Path(args.output),
        sheet_arg=parsed_sheet,
    )


if __name__ == "__main__":
    main()
